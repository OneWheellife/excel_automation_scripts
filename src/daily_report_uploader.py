import pandas as pd
from datetime import datetime, timedelta
import sys
import os
import traceback
import openpyxl  # openpyxl も引き続き利用
import shutil
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# win32com をインポート
try:
    import win32com.client
    import pythoncom
except ImportError:
    print("エラー: このスクリプトの実行には pywin32 ライブラリが必要です。")
    print("コマンドプロンプトで `pip install pywin32` を実行してインストールしてください。")
    sys.exit(1)

# --- win32com ヘルパー関数 ---


def get_excel_colors_win32(file_path, sheet_name, map_range_str="B1:G2", data_start_row=4, max_row=None, max_col=None):
    """
    win32comを使用してExcelファイルから指定範囲の色情報を取得する関数。

    Args:
        file_path (str): Excelファイルのパス。
        sheet_name (str): シート名。
        map_range_str (str): 色マッピングを取得する範囲 (例: "B1:G2")。
        data_start_row (int): データが開始する行番号 (1-based)。
        max_row (int, optional): 色を取得する最大行番号。指定がない場合はシート全体を試みるが非推奨。
        max_col (int, optional): 色を取得する最大列番号。指定がない場合はシート全体を試みるが非推奨。

    Returns:
        tuple: (color_map, cell_rgb_map)
               color_map (dict): 背景色RGBタプルをキー、セル値を値とする辞書。
               cell_rgb_map (dict): セル座標タプル(row, col)をキー、背景色RGBタプルを値とする辞書。
               エラー時は (None, None) を返す。
    """
    color_map = {}
    cell_rgb_map = {}
    excel = None
    workbook = None
    pythoncom.CoInitialize()  # COMライブラリ初期化

    def bgr_to_rgb(bgr_int):
        """BGR整数値をRGBタプルに変換"""
        if bgr_int is None:
            return None
        try:
            bgr_int = int(bgr_int)
            # 0x00FFFFFF (白) までの値かチェック
            if 0 <= bgr_int <= 16777215:
                b = bgr_int & 255
                g = (bgr_int >> 8) & 255
                r = (bgr_int >> 16) & 255
                return (r, g, b)
            else:
                return None  # 範囲外の値は無効とする
        except (ValueError, TypeError, OverflowError):
            return None

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # バックグラウンドで実行
        excel.DisplayAlerts = False  # 警告を非表示

        # 絶対パスに変換してファイルを開く
        abs_file_path = os.path.abspath(file_path)
        try:
            workbook = excel.Workbooks.Open(abs_file_path)
        except pythoncom.com_error as e:
            print(f"エラー: Excelファイルのオープンに失敗しました: {abs_file_path}")
            print(f"詳細: {e}")
            return None, None

        try:
            sheet = workbook.Sheets(sheet_name)
            # openpyxlで事前に取得した最大行・列がない場合はシートから取得（ただし重い可能性）
            if max_row is None:
                max_row = sheet.UsedRange.Rows.Count + sheet.UsedRange.Row - 1
                print(f"警告: 最大行をシートから推定しました ({max_row})。時間がかかる場合があります。")
            if max_col is None:
                max_col = sheet.UsedRange.Columns.Count + sheet.UsedRange.Column - 1
                print(f"警告: 最大列をシートから推定しました ({max_col})。時間がかかる場合があります。")

        except pythoncom.com_error:
            print(f"エラー: シート '{sheet_name}' が見つかりません。")
            workbook.Close(SaveChanges=False)
            excel.Quit()
            return None, None

        print(
            f"win32com: シート '{sheet_name}' の色情報を読み込み中 (範囲: R{data_start_row}-R{max_row}, C1-C{max_col})...")

        # 1. color_map の作成 (指定範囲: B1:G2 など)
        try:
            map_range = sheet.Range(map_range_str)
            for cell in map_range:
                cell_value = str(cell.Value) if cell.Value is not None else ""
                rgb = None
                # ColorIndex が -4142 (xlColorIndexNone) でない場合のみ色を取得
                if cell.Interior.ColorIndex != -4142:
                    rgb = bgr_to_rgb(cell.Interior.Color)

                if rgb and cell_value:
                    color_map[rgb] = cell_value
        except pythoncom.com_error as e_map:
            print(f"警告: 色マッピング範囲 '{map_range_str}' の処理中にエラー: {e_map}")
            # エラーでも処理は続行

        # 2. cell_rgb_map の作成 (データ範囲)
        # パフォーマンスのため、範囲を限定することが望ましい
        if max_row > 20000 or max_col > 100:  # 仮の上限設定
            print(f"警告: 処理範囲が広すぎます (行:{max_row}, 列:{max_col})。")
            print("      範囲を限定するか、ファイル内容を確認してください。中断します。")
            workbook.Close(SaveChanges=False)
            excel.Quit()
            return None, None

        # セルごとに色を取得
        for r in range(data_start_row, max_row + 1):
            for c in range(1, max_col + 1):
                try:
                    cell = sheet.Cells(r, c)
                    # ColorIndexチェックで高速化
                    if cell.Interior.ColorIndex != -4142:
                        rgb = bgr_to_rgb(cell.Interior.Color)
                        if rgb:
                            cell_rgb_map[(r, c)] = rgb
                except pythoncom.com_error:
                    pass  # セルアクセスエラーは無視して続行
                except Exception as e_cell:
                    pass  # その他のエラーも無視して続行

        print(
            f"win32com: 色情報の読み込み完了 (マップ:{len(color_map)}件, セル色:{len(cell_rgb_map)}件)")

        # Excelを閉じる
        workbook.Close(SaveChanges=False)
        excel.Quit()

        return color_map, cell_rgb_map

    except pythoncom.com_error as e:
        print(f"Excel操作(win32com)中にCOMエラーが発生しました: {e}")
        traceback.print_exc()
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass
        return None, None
    except Exception as e:
        print(f"win32comでの色情報取得中に予期せぬエラーが発生しました: {e}")
        traceback.print_exc()
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass
        return None, None
    finally:
        # オブジェクト参照を解放（念のため）
        sheet = None
        workbook = None
        excel = None
        pythoncom.CoUninitialize()  # COMライブラリ終了処理


# --- メイン処理関数 ---
def process_report():
    """
    ユーザーにExcelファイルを選択させ、指定された条件でデータを抽出し、
    '報告' と '保留' のシートに分けて書式設定して保存する関数。
    確認した人情報は win32com を使用して取得したセルの色に基づいて追加する。
    *** この関数は Windows + Excel 環境が必要です ***
    """
    # 個人情報・社名を含まない公開用ディレクトリ名に変更
    target_dir = r"C:\ExcelData\Project_ReportList"

    if not os.path.isdir(target_dir):
        sys.exit(f"エラー: 指定されたディレクトリが見つかりません: {target_dir}")

    excel_files = []
    try:
        for filename in os.listdir(target_dir):
            if filename.lower().endswith((".xlsx", ".xls")):
                full_path = os.path.join(target_dir, filename)
                if os.path.isfile(full_path):
                    excel_files.append(full_path)
    except Exception as e:
        sys.exit(f"エラー: ディレクトリ内のファイルリスト取得中にエラーが発生しました: {target_dir}\n{e}")

    if not excel_files:
        sys.exit(
            f"エラー: 指定されたディレクトリに処理対象のExcelファイル (.xlsx または .xls) が見つかりません: {target_dir}")

    # 更新日時が最新のファイルを選択
    try:
        latest_file = max(excel_files, key=os.path.getmtime)
        file_path = latest_file
        print(f"処理対象ファイル (最新): {file_path}")
    except Exception as e:
        sys.exit(f"エラー: 最新ファイルの特定中にエラーが発生しました。\n{e}")

    # ファイル形式チェックは win32com 前提なら緩和してもよいが、pandas用に残す
    file_extension = os.path.splitext(file_path)[1].lower()
    engine_pd = 'openpyxl' if file_extension == '.xlsx' else 'xlrd' if file_extension == '.xls' else None
    if engine_pd is None:
        print(f"警告: pandas がサポートしない可能性のあるファイル形式です: {file_extension}")

    table_name = "Sheet1"  # 元データのシート名
    required_cols = ["報告内容", "報告内容2", "保留案件"]  # 必須の列名

    # --- 色情報取得 (win32com を使用) ---
    color_map = None
    cell_rgb_map = None
    try:
        print("win32com: Excelから色情報を取得します...")
        # openpyxlで事前に最大行・列を取得（win32comの処理範囲特定のため）
        max_row_check = 0
        max_col_check = 0
        try:
            wb_check = openpyxl.load_workbook(file_path, read_only=True)
            if table_name in wb_check.sheetnames:
                sheet_check = wb_check[table_name]
                max_row_check = sheet_check.max_row
                max_col_check = sheet_check.max_column
            else:
                print(f"エラー: openpyxlでシート '{table_name}' が見つかりません。")
                sys.exit(f"シート '{table_name}' がファイル内に見つかりません。")
            wb_check.close()
            if max_row_check == 0 or max_col_check == 0:
                print(f"警告: openpyxlでシート '{table_name}' の有効な範囲を取得できませんでした。")
                max_row_check = None
                max_col_check = None

        except Exception as e_openpyxl_check:
            print(f"警告: openpyxlでの範囲確認中にエラー: {e_openpyxl_check}")
            print("      win32comに範囲推定を試みさせます。")
            max_row_check = None
            max_col_check = None

        # win32com で色情報を取得
        color_map, cell_rgb_map = get_excel_colors_win32(
            file_path, table_name, "B1:G2", 4, max_row_check, max_col_check
        )

        if color_map is None or cell_rgb_map is None:
            sys.exit("Excelからの色情報の取得に失敗しました。処理を中断します。")
        if not color_map:
            print(
                f"警告: シート '{table_name}' の B1:G2 範囲に、色と文字列のマッピングが見つかりませんでした。")

    except Exception as e_color_main:
        print(f"エラー: 色情報取得処理の呼び出し中にエラーが発生しました: {e_color_main}")
        traceback.print_exc()
        sys.exit()

    try:
        # --- データ読み込み (pandas) ---
        print("pandas: Excelからデータを読み込みます...")
        df_all = pd.read_excel(
            file_path, sheet_name=table_name, header=2, engine=engine_pd)

        # A列が空でない行をフィルタリング (元のインデックスを保持)
        df = df_all[df_all.iloc[:, 0].notna()].copy()
        if df.empty:
            print("A列にデータが含まれる行が見つかりませんでした。ファイルは作成されません。")
            sys.exit()

        # 必要な列の存在確認
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            sys.exit(
                f"エラー: 必要な列が見つかりません: {', '.join(missing_cols)}。シート '{table_name}' の3行目のヘッダーを確認してください。")

        # --- データ抽出 (報告・保留) ---
        horyu_condition = df["保留案件"].notna()
        df_horyu = df[horyu_condition].copy()

        df_not_horyu = df[~horyu_condition].copy()
        hokoku_condition = df_not_horyu["報告内容"].notna(
        ) | df_not_horyu["報告内容2"].notna()
        df_hokoku = df_not_horyu[hokoku_condition].copy()
        print(f"データ抽出完了: 報告 {len(df_hokoku)}件, 保留 {len(df_horyu)}件")

        # --- 確認した人情報追加 (取得済みの色情報を使用) ---
        assignees_dict = {}  # df のインデックスをキーとして確認した人を格納
        if not df.empty and color_map:
            print("確認した人情報を付与しています...")
            skip_col_indices = {5, 6, 7}  # F, G, H 列 (0-based)
            excel_row_offset = 4  # df.index=0 は Excel 4行目

            for pd_index in df.index:
                excel_row_index = pd_index + excel_row_offset
                assignee = None

                for col_idx in range(len(df.columns)):
                    if col_idx in skip_col_indices:
                        continue

                    excel_coord = (excel_row_index, col_idx + 1)
                    cell_rgb = cell_rgb_map.get(excel_coord)

                    if cell_rgb and cell_rgb in color_map:
                        assignee = color_map[cell_rgb]
                        break

                if assignee:
                    assignees_dict[pd_index] = assignee

            if not df_hokoku.empty:
                df_hokoku['確認した人'] = df_hokoku.index.map(
                    assignees_dict).fillna('')
            if not df_horyu.empty:
                df_horyu['確認した人'] = df_horyu.index.map(
                    assignees_dict).fillna('')
            print("確認した人情報の付与完了。")

        elif not color_map:
            print("確認した人情報の追加スキップ: 色マッピングが空です。")
            if not df_hokoku.empty:
                df_hokoku['確認した人'] = ''
            if not df_horyu.empty:
                df_horyu['確認した人'] = ''
        else:
            pass

        # --- 出力ファイル処理 ---
        today_str = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        output_filename = f"{today_str}_プロジェクト報告.xlsx"
        output_dir = os.path.dirname(file_path)
        output_path = os.path.join(output_dir, output_filename)

        # ExcelWriterで複数シートに書き込み
        if not df_hokoku.empty or not df_horyu.empty:
            print(f"処理結果を '{output_path}' に書き込みます...")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                if not df_hokoku.empty:
                    if '確認した人' in df_hokoku.columns:
                        cols_hokoku = [
                            col for col in df_hokoku if col != '確認した人'] + ['確認した人']
                        df_hokoku[cols_hokoku].to_excel(
                            writer, sheet_name='報告', index=False)
                    else:
                        df_hokoku.to_excel(
                            writer, sheet_name='報告', index=False)

                if not df_horyu.empty:
                    if '確認した人' in df_horyu.columns:
                        cols_horyu = [
                            col for col in df_horyu if col != '確認した人'] + ['確認した人']
                        df_horyu[cols_horyu].to_excel(
                            writer, sheet_name='保留', index=False)
                    else:
                        df_horyu.to_excel(writer, sheet_name='保留', index=False)

            try:
                print("出力ファイルの書式を設定します...")
                workbook = openpyxl.load_workbook(output_path)

                common_table_style_args = {
                    "showFirstColumn": False, "showLastColumn": False,
                    "showRowStripes": True, "showColumnStripes": False
                }
                table_names_to_remove = []

                if '報告' in workbook.sheetnames and workbook['報告'].max_row > 1:
                    ws_hokoku = workbook['報告']
                    table_range_hokoku = f"A1:{get_column_letter(ws_hokoku.max_column)}{ws_hokoku.max_row}"
                    tab_hokoku = Table(displayName="Table_報告",
                                       ref=table_range_hokoku)
                    style_hokoku = TableStyleInfo(
                        name="TableStyleMedium2", **common_table_style_args)
                    tab_hokoku.tableStyleInfo = style_hokoku
                    if "Table_報告" in ws_hokoku.tables:
                        table_names_to_remove.append(("報告", "Table_報告"))
                    ws_hokoku.add_table(tab_hokoku)

                if '保留' in workbook.sheetnames and workbook['保留'].max_row > 1:
                    ws_horyu = workbook['保留']
                    table_range_horyu = f"A1:{get_column_letter(ws_horyu.max_column)}{ws_horyu.max_row}"
                    tab_horyu = Table(displayName="Table_保留",
                                      ref=table_range_horyu)
                    style_horyu = TableStyleInfo(
                        name="TableStyleMedium4", **common_table_style_args)
                    tab_horyu.tableStyleInfo = style_horyu
                    if "Table_保留" in ws_horyu.tables:
                        table_names_to_remove.append(("保留", "Table_保留"))
                    ws_horyu.add_table(tab_horyu)

                meiryo_font = Font(name='Meiryo UI', size=11)

                for sheet_name in ['報告', '保留']:
                    if sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        if ws.max_row <= 1:
                            continue

                        for row in ws.iter_rows(min_row=1):
                            for cell in row:
                                cell.font = meiryo_font

                        for col in ws.columns:
                            max_length = 0
                            column_letter = col[0].column_letter

                            header_cell = ws.cell(row=1, column=col[0].column)
                            header_len = 0
                            if header_cell.value:
                                try:
                                    for char in str(header_cell.value):
                                        header_len += 1.8 if ord(
                                            char) > 255 else 1
                                except:
                                    pass
                            max_length = header_len

                            for cell in col:
                                if cell.row == 1:
                                    continue
                                try:
                                    if cell.value is not None:
                                        cell_value_str = str(cell.value)
                                        current_length = 0
                                        for char in cell_value_str:
                                            current_length += 1.8 if ord(
                                                char) > 255 else 1
                                        if current_length > max_length:
                                            max_length = current_length
                                except Exception:
                                    pass

                            adjusted_width = max(max_length * 1.1 + 1, 10)
                            adjusted_width = min(adjusted_width, 60)
                            ws.column_dimensions[column_letter].width = adjusted_width

                workbook.save(output_path)
                print(f"処理完了: '{output_path}' に結果を出力しました。")

                # ファイルを移動
                destination_dir = r"C:\ExcelData\Project_ReportList"
                try:
                    os.makedirs(destination_dir, exist_ok=True)
                    destination_path = os.path.join(
                        destination_dir, output_filename)
                    shutil.move(output_path, destination_path)
                    print(f"ファイルを '{destination_path}' に移動しました。")
                except FileNotFoundError:
                    print(f"エラー: 移動元のファイルが見つかりません: {output_path}")
                except PermissionError:
                    print(f"エラー: 移動先に書き込む権限がありません: {destination_dir}")
                except Exception as e_move:
                    print(f"エラー: ファイル移動中に予期せぬエラーが発生しました: {e_move}")
                    traceback.print_exc()

            except Exception as e_format:
                print(f"エラー: 出力ファイルの書式設定中にエラーが発生しました: {e_format}")
                traceback.print_exc()
                print(f"注意: 書式設定前のファイルが '{output_path}' に保存されている可能性があります。")

        else:
            print("抽出対象となるデータが見つからなかったため、ファイルは作成されませんでした。")

    except FileNotFoundError:
        sys.exit(f"エラー: 指定されたファイルが見つかりません: {file_path}")
    except ValueError as e:
        print(f"エラー: ファイル処理中にValueErrorが発生しました: {e}")
        if f"Worksheet named '{table_name}' not found" in str(e):
            print(f"原因: シート名 '{table_name}' が見つかりません。ファイル内のシート名を確認してください。")
        elif "support for xls files" in str(e):
            print("原因: .xls ファイルの読み込みに必要な 'xlrd' ライブラリ関連のエラーの可能性があります。")
        traceback.print_exc()
        sys.exit()
    except ImportError as e:
        print(f"エラー: 必要なライブラリが見つかりません: {e}")
        if 'xlrd' in str(e):
            print("'.xls' ファイル処理には 'xlrd' が必要です。`pip install xlrd` でインストールしてください。")
        elif 'openpyxl' in str(e):
            print(
                "'.xlsx' ファイル処理には 'openpyxl' が必要です。`pip install openpyxl` でインストールしてください。")
        sys.exit()
    except KeyError as e:
        print(
            f"エラー: 必要な列名が見つかりません: {e}。シート '{table_name}' の3行目のヘッダーを確認してください。")
        sys.exit()
    except Exception as e:
        print(f"予期せぬエラーが発生しました: {e}")
        traceback.print_exc()
        sys.exit()


if __name__ == "__main__":
    print("--- プロジェクト報告処理 (win32com 色情報取得版) ---")
    process_report()
    print("--- 処理終了 ---")
