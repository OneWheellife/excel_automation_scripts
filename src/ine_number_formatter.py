import openpyxl
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Border, Side # 未使用のPatternFill, Font, Alignmentを削除
# from openpyxl.utils import get_column_letter # 未使用のため削除
import os
import sys # sysモジュールを追加
import time # timeモジュールをトップレベルでインポート

# pywin32のインポート試行
try:
    import win32com.client
    import pythoncom # COM初期化に必要
except ImportError:
    messagebox.showerror(
        "ライブラリ不足エラー",
        "この機能には 'pywin32' ライブラリが必要です。\n"
        "コマンドプロンプトまたはターミナルで以下のコマンドを実行してインストールしてください:\n\n"
        "pip install pywin32"
    )
    sys.exit(1) # スクリプトを終了

# --- 定数 ---
PRIMARY_NAME = "HOUSES.BUILD_START"
SECONDARY_NAME = "HOUSES.SCHEDULE_DATE"
TARGET_SHEET_KEYWORD = "L線番表"
MAX_ROWS_TO_PROCESS = 1000
EXCLUDED_ROWS_F = range(8, 13) # F列の更新を除外する行 (1-based)

# --- 罫線スタイル定義 (Sideオブジェクト) ---
side_thin = Side(style='thin')
side_hair = Side(style='hair') # 内部横線用 (VBAのxlHairline)
side_none = Side(style=None) # 罫線なし

# --- 正規表現パターン ---
# パターン: (2桁数字 + 空白* + HUB) または (HUB + (空白 or -)* + 2桁数字)
hub_pattern = re.compile(r"(\d{2})\s*HUB|HUB[\s-]*(\d{2})", re.IGNORECASE)

def get_named_range_value(workbook, sheet, name):
    """指定された名前付き範囲の値をシートレベルまたはブックレベルで検索して返す"""
    try:
        # シートレベルで検索
        defined_name = workbook.defined_names[name]
        # ローカルスコープ（シートレベル）の名前かチェック
        if defined_name.localSheetId is not None and defined_name.localSheetId == workbook.worksheets.index(sheet):
            # RefersToRange形式 ($A$1 など) を想定
            dest = defined_name.destinations
            if dest:
                for sheetname, cell_coord in dest:
                    if sheetname == sheet.title:
                        return sheet[cell_coord].value
        # ブックレベルの名前もチェック (localSheetId is None)
        elif defined_name.localSheetId is None:
             dest = defined_name.destinations
             if dest:
                # ブックレベルの場合、最初の宛先を返す (VBAの挙動に近い)
                sheetname, cell_coord = next(iter(dest))
                # 参照先が現在のシートか確認した方が良い場合もあるが、
                # VBAではブックレベルならどのシートからでもアクセスできるため、
                # openpyxlのworkbook[sheetname][cell_coord]で値を取得
                return workbook[sheetname][cell_coord].value

    except KeyError:
        pass # シートレベルで見つからない場合はブックレベルへ

    # ブックレベルで検索 (シートレベルで見つからないか、ブックレベルの名前だった場合)
    try:
        defined_name = workbook.defined_names[name]
        if defined_name.localSheetId is None: # ブックレベルの名前か確認
            dest = defined_name.destinations
            if dest:
                sheetname, cell_coord = next(iter(dest))
                return workbook[sheetname][cell_coord].value
    except KeyError:
        pass # ブックレベルでも見つからない

    return None # 見つからなかった場合

def close_excel_if_open(filepath_to_check):
    """指定されたファイルパスをExcelが開いているか確認し、開いていれば閉じる"""
    try:
        pythoncom.CoInitialize() # COMライブラリを初期化
        excel = win32com.client.Dispatch("Excel.Application")
        # 実行中のExcelインスタンスを取得しようとする (GetActiveObjectは不安定な場合があるためDispatchを使用)
        # 注意: これだと新しいインスタンスが起動してしまう可能性がある。
        # 実行中のインスタンスを取得するにはGetActiveObjectが一般的だが、エラー処理が必要。

        # より安定した方法: 実行中の全Excelプロセスをチェックする
        found_and_closed = False
        abs_filepath_to_check = os.path.abspath(filepath_to_check)

        # ROT (Running Object Table) からExcelインスタンスを取得
        ctx = pythoncom.CreateBindCtx(0)
        running_objects = pythoncom.GetRunningObjectTable()
        monikers = running_objects.EnumRunning()

        for moniker in monikers:
            try:
                obj_name = moniker.GetDisplayName(ctx, None)
                # Excelファイルを開いているインスタンスか確認
                if "Excel.Sheet" in obj_name or abs_filepath_to_check in obj_name: # ファイル名が含まれるかチェック
                    obj = running_objects.GetObject(moniker)
                    # IDispatchインターフェースを取得してExcelアプリケーションオブジェクトにアクセス
                    excel_app = obj.QueryInterface(pythoncom.IID_IDispatch)
                    # Excel.Applicationオブジェクトか確認 (より確実な方法)
                    # excel_app = win32com.client.Dispatch(excel_app) # これでApplicationオブジェクトに変換できるか？

                    # 開いているブックを確認 (より確実な方法)
                    # GetObjectで取得したオブジェクトが直接Applicationとは限らないため、
                    # Applicationプロパティ経由でアクセスする
                    try:
                        excel_instance = excel_app.Application # Applicationオブジェクトを取得
                        for wb in excel_instance.Workbooks:
                            try:
                                wb_fullpath = os.path.abspath(wb.FullName)
                                if wb_fullpath == abs_filepath_to_check:
                                    print(f"情報: ファイル '{os.path.basename(filepath_to_check)}' を開いているExcelプロセスが見つかりました。閉じます...")
                                    # wb.Close(SaveChanges=False) # ワークブックだけ閉じる (安全だが、アプリが残る)
                                    excel_instance.Quit() # アプリケーションごと閉じる (リスクあり)
                                    found_and_closed = True
                                    print(f"情報: Excelプロセスを閉じました。")
                                    # 重要: プロセス終了後、少し待機時間を設ける
                                    # time モジュールはトップレベルでインポート済み
                                    time.sleep(2) # 2秒待機
                                    break # 対象を見つけて閉じたので内部ループを抜ける
                            except Exception as e_wb:
                                print(f"警告: ワークブック '{getattr(wb, 'Name', '不明')}' のパス取得または比較中にエラー: {e_wb}")
                                continue # 次のワークブックへ
                        if found_and_closed:
                            break # 対象を見つけて閉じたので外部ループを抜ける
                    except AttributeError:
                         # Applicationプロパティがないオブジェクトだった場合 (例: Workbookオブジェクト直接など)
                         # このオブジェクトが目的のファイルを開いているか確認
                         try:
                             if hasattr(excel_app, 'FullName') and os.path.abspath(excel_app.FullName) == abs_filepath_to_check:
                                 print(f"情報: ファイル '{os.path.basename(filepath_to_check)}' を開いているExcelオブジェクトが見つかりました。関連プロセスを閉じます...")
                                 # このオブジェクトに関連するApplicationを取得して閉じる
                                 app_to_quit = excel_app.Application
                                 app_to_quit.Quit()
                                 found_and_closed = True
                                 print(f"情報: Excelプロセスを閉じました。")
                                 # time モジュールはトップレベルでインポート済み
                                 time.sleep(2)
                                 break
                         except Exception as e_obj:
                             print(f"警告: Excelオブジェクトの確認中にエラー: {e_obj}")
                             continue
                    except Exception as e_app:
                        print(f"警告: Excelアプリケーションの処理中にエラー: {e_app}")
                        continue # 次のモニカへ

            except Exception as e_moniker:
                # print(f"デバッグ: モニカ処理中にエラー: {e_moniker}") # デバッグ用
                continue # 次のモニカへ

        if not found_and_closed:
            print(f"情報: ファイル '{os.path.basename(filepath_to_check)}' はExcelで開かれていないようです。")

    except Exception as e:
        messagebox.showwarning("Excelプロセス確認エラー", f"実行中のExcelプロセスを確認または終了する際にエラーが発生しました。\n手動でファイルが閉じていることを確認してください。\n\n詳細: {e}")
    finally:
        pythoncom.CoUninitialize() # COMライブラリを解放

def set_borders_for_range(sheet, start_row, end_row):
    """指定された範囲に罫線を設定する (再修正版: 初期化と4辺の確実な設定)"""
    if start_row <= 0 or end_row < start_row:
        return

    target_range_str = f"A{start_row}:I{end_row}"
    min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(target_range_str)

    # 範囲内のすべてのセルをループし、各セルの4辺を決定して設定
    for r_idx in range(min_row, max_row + 1):
        for c_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=r_idx, column=c_idx)

            # 各辺のスタイルを決定
            left_style = side_none
            right_style = side_none
            top_style = side_none
            bottom_style = side_none

            # 上辺
            if r_idx == min_row:
                top_style = side_thin  # 範囲の最上辺
            else:
                top_style = side_hair  # 内部横線 (上のセルの下辺に相当)

            # 下辺
            if r_idx == max_row:
                bottom_style = side_thin # 範囲の最下辺
            # else: # 内部の下辺は、下のセルの上辺(hair)で描画されるのでNoneのまま -> この考え方は間違い
            # 各セルの下辺も設定する必要がある
            elif r_idx < max_row: # 最下行でない場合
                 bottom_style = side_hair # 内部横線

            # 左辺
            if c_idx == min_col:
                left_style = side_thin  # 範囲の最左辺
            else:
                left_style = side_thin  # 内部縦線 (左のセルの右辺に相当)

            # 右辺
            if c_idx == max_col:
                right_style = side_thin # 範囲の最右辺
            # else: # 内部の右辺は、右のセルの左辺(thin)で描画されるのでNoneのまま -> この考え方は間違い
            # 各セルの右辺も設定する必要がある
            elif c_idx < max_col: # 最右列でない場合
                right_style = side_thin # 内部縦線

            # Borderオブジェクトを作成して適用
            cell.border = Border(left=left_style, right=right_style, top=top_style, bottom=bottom_style)


def process_report_sheets(filepath):
    """メインの処理関数"""
    # --- ファイルを開く前に、Excelで開かれていれば閉じる ---
    close_excel_if_open(filepath)
    # -------------------------------------------------

    try:
        # 少し待機してからファイルを開く
        time.sleep(0.5) # Excel終了待ちの後、ファイルハンドル解放のために少し待つ
        workbook = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        messagebox.showerror("エラー", f"ファイルが見つかりません:\n{filepath}")
        return
    except Exception as e: # PermissionErrorなどもここで捕捉される可能性がある
        error_detail = str(e)
        if "Permission denied" in error_detail:
             messagebox.showerror("エラー", f"ファイルへのアクセスが拒否されました:\n{filepath}\n\nファイルがExcelなどの他のプログラムで開かれている可能性があります。\nファイルを閉じてから再度実行してください。\n\n詳細: {e}")
        else:
            messagebox.showerror("エラー", f"ファイルを開けませんでした:\n{filepath}\n\n詳細: {e}")
        return

    value_to_use = None
    found_value = False

    # 1. HOUSES.BUILD_START の値を取得 (ブックレベル優先)
    try:
        # まずブックレベルで試す
        defined_name = workbook.defined_names[PRIMARY_NAME]
        if defined_name.localSheetId is None: # ブックレベルの名前か確認
            dest = defined_name.destinations
            if dest:
                sheetname, cell_coord = next(iter(dest))
                value_to_use = workbook[sheetname][cell_coord].value
                if value_to_use is not None and value_to_use != "":
                    found_value = True
    except KeyError:
        pass # ブックレベルで見つからない場合は次へ

    # 2. HOUSES.BUILD_START の値を取得 (シートレベル、ブックレベルで見つからない場合)
    if not found_value:
        for sheet in workbook.worksheets:
            try:
                defined_name = workbook.defined_names[PRIMARY_NAME]
                # ローカルスコープ（シートレベル）の名前かチェック
                if defined_name.localSheetId is not None and defined_name.localSheetId == workbook.worksheets.index(sheet):
                    dest = defined_name.destinations
                    if dest:
                        for sheetname, cell_coord in dest:
                            if sheetname == sheet.title:
                                value_to_use = sheet[cell_coord].value
                                if value_to_use is not None and value_to_use != "":
                                    found_value = True
                                    break # 見つかったらループを抜ける
                if found_value:
                    break
            except KeyError:
                continue # このシートにはその名前がない

    # 3. HOUSES.BUILD_START が見つからなかった場合、HOUSES.SCHEDULE_DATE を試す
    if not found_value:
        secondary_value = None
        found_secondary = False
        # 3.1 ブックレベルで試す
        try:
            defined_name = workbook.defined_names[SECONDARY_NAME]
            if defined_name.localSheetId is None:
                dest = defined_name.destinations
                if dest:
                    sheetname, cell_coord = next(iter(dest))
                    secondary_value = workbook[sheetname][cell_coord].value
                    if secondary_value is not None and secondary_value != "":
                        found_secondary = True
        except KeyError:
            pass

        # 3.2 シートレベルで試す
        if not found_secondary:
            for sheet in workbook.worksheets:
                try:
                    defined_name = workbook.defined_names[SECONDARY_NAME]
                    if defined_name.localSheetId is not None and defined_name.localSheetId == workbook.worksheets.index(sheet):
                        dest = defined_name.destinations
                        if dest:
                            for sheetname, cell_coord in dest:
                                if sheetname == sheet.title:
                                    secondary_value = sheet[cell_coord].value
                                    if secondary_value is not None and secondary_value != "":
                                        found_secondary = True
                                        break
                    if found_secondary:
                        break
                except KeyError:
                    continue

        if found_secondary:
            value_to_use = secondary_value
            found_value = True
        else:
            messagebox.showerror("エラー", f"名前付きセル '{PRIMARY_NAME}' および '{SECONDARY_NAME}' が見つからないか、有効な値を持っていません。")
            return # 処理中断

    # --- ここから先の処理では value_to_use を使用 ---
    if value_to_use is None: # 念のためチェック
         messagebox.showerror("エラー", "F列に入力するための値を取得できませんでした。")
         return

    # 4. 名前に "L線番表" を含み、B6セルが空でないシートを検索して取得
    target_sheets = []
    for sheet in workbook.worksheets:
        if TARGET_SHEET_KEYWORD in sheet.title:
            b6_value = sheet["B6"].value
            if b6_value is not None and b6_value != "":
                target_sheets.append(sheet)

    if not target_sheets:
        messagebox.showerror("エラー", f"名前に '{TARGET_SHEET_KEYWORD}' を含み、かつB6セルに値があるシートが見つかりません。")
        return
    # VBA版では複数シートが見つかっても確認なしで処理していたので、Python版でもそのまま処理

    # 5. 対象シートの処理
    all_results = [] # 全シートの結果を格納

    for sheet in target_sheets:
        f_updated_address = []
        d_gokaku_added_address = []
        b_hub_replaced_address = []
        b_ip_replaced_address = []
        d_hub_gokaku_added_address = []
        updated_count = 0
        any_updates = False

        border_start_row = 0
        border_end_row = 0

        # 1行目から指定行数までを処理
        for i in range(1, MAX_ROWS_TO_PROCESS + 1):
            # セルオブジェクト取得
            a_cell = sheet.cell(row=i, column=1) # A列
            b_cell = sheet.cell(row=i, column=2) # B列
            d_cell = sheet.cell(row=i, column=4) # D列
            f_cell = sheet.cell(row=i, column=6) # F列

            # 値を取得 (文字列として扱い、前後の空白を削除)
            # .valueがNoneの場合も考慮してstr()で変換
            a_value = a_cell.value # 罫線判定用なので型はそのまま
            b_value_str = str(b_cell.value).strip() if b_cell.value is not None else ""
            d_value_str = str(d_cell.value).strip() if d_cell.value is not None else ""
            f_value = f_cell.value # F列は空かどうかを直接判定

            f_cell_is_empty = f_value is None or f_value == ""
            d_cell_is_empty = d_value_str == "" # Trimして空文字なら空とみなす

            # --- 条件1: D列が "合格" で F列が空の場合 ---
            if d_value_str.lower() == "合格" and f_cell_is_empty:
                if i not in EXCLUDED_ROWS_F:
                    f_cell.value = value_to_use
                    updated_count += 1
                    f_updated_address.append(f_cell.coordinate)
                    any_updates = True

            # --- 条件2: D列が空白で F列が空の場合 ---
            elif d_cell_is_empty and f_cell_is_empty:
                # B列が3文字以下の空でない文字列で、かつ "ONU" や "HUB" でない場合
                if 0 < len(b_value_str) <= 3:
                    b_value_upper = b_value_str.upper()
                    if b_value_upper != "ONU" and b_value_upper != "HUB":
                        if i not in EXCLUDED_ROWS_F:
                            f_cell.value = value_to_use
                            d_cell.value = "合格"
                            updated_count += 2 # FとDの2箇所更新
                            f_updated_address.append(f_cell.coordinate)
                            d_gokaku_added_address.append(d_cell.coordinate)
                            any_updates = True

            # --- HUB パターンの置換 ---
            match = hub_pattern.search(b_value_str)
            if match:
                num_part = ""
                # マッチした部分から数字を取得 (group(1) or group(2))
                if match.group(1):
                    num_part = match.group(1)
                elif match.group(2):
                    num_part = match.group(2)

                # numPart が2桁の数字の場合のみ処理
                if len(num_part) == 2 and num_part.isdigit():
                    last_digit = num_part[-1]
                    new_value = f"HUB-{last_digit}"
                    if str(b_cell.value) != new_value: # 値が変わる場合のみ更新
                        b_cell.value = new_value
                        updated_count += 1
                        b_hub_replaced_address.append(b_cell.coordinate)
                        any_updates = True

            # --- IPアドレス置換 ("10.32.0.1" -> "10.128.0.1") ---
            if b_value_str == "10.32.0.1":
                 if str(b_cell.value) != "10.128.0.1": # 値が変わる場合のみ更新
                    b_cell.value = "10.128.0.1"
                    updated_count += 1
                    b_ip_replaced_address.append(b_cell.coordinate)
                    any_updates = True

            # --- B列に "HUB" が含まれ、かつD列に "合格" が含まれていない場合、D列に "合格" を追記 ---
            # 注意: HUBパターン置換後の値で判定すべきか、元の値で判定すべきか？
            # VBAでは置換前の bValue で判定しているので、それに合わせる
            if "HUB" in b_value_str.upper(): # 大文字小文字区別せず
                if "合格" not in d_value_str.lower(): # 大文字小文字区別せず
                    if str(d_cell.value).lower() != "合格": # 値が"合格"でない場合のみ更新
                        d_cell.value = "合格"
                        updated_count += 1
                        d_hub_gokaku_added_address.append(d_cell.coordinate)
                        any_updates = True

            # --- 罫線対象行か判定 ---
            is_target_row = False
            # A列が "G" で始まるかチェック
            if isinstance(a_value, str) and a_value.strip().upper().startswith("G"):
                is_target_row = True
            # A列が 1 から 99 までの整数かチェック
            elif isinstance(a_value, (int, float)) and 1 <= a_value <= 99 and int(a_value) == a_value:
                 is_target_row = True

            # --- 罫線対象ブロックの開始行と終了行を管理 ---
            if is_target_row:
                if border_start_row == 0:
                    border_start_row = i # 新しいブロックの開始
                border_end_row = i # ブロックの終了行を更新
            else:
                # 対象行でない場合 (またはブロックが終わった場合)
                if border_start_row > 0:
                    # 直前まで対象ブロックが続いていたので、罫線を設定
                    set_borders_for_range(sheet, border_start_row, border_end_row)
                    # ブロック情報をリセット
                    border_start_row = 0
                    border_end_row = 0

        # --- ループ終了後、最後のブロックが残っている可能性があるので処理 ---
        if border_start_row > 0:
             set_borders_for_range(sheet, border_start_row, border_end_row)

        # --- シートごとの結果メッセージ作成 ---
        result_msg = f"シート '{sheet.title}' の処理結果:\n\n"
        if any_updates:
            if f_updated_address:
                result_msg += f"F列に日付を入力 ({len(f_updated_address)}件): {', '.join(f_updated_address)}\n"
            if d_gokaku_added_address:
                result_msg += f"D列に'合格'を入力 (条件2, {len(d_gokaku_added_address)}件): {', '.join(d_gokaku_added_address)}\n"
            if b_hub_replaced_address:
                result_msg += f"B列のHUBパターンを置換 ({len(b_hub_replaced_address)}件): {', '.join(b_hub_replaced_address)}\n"
            if b_ip_replaced_address:
                result_msg += f"B列のIPアドレスを置換 ({len(b_ip_replaced_address)}件): {', '.join(b_ip_replaced_address)}\n"
            if d_hub_gokaku_added_address:
                result_msg += f"D列に'合格'を追記 (HUB行, {len(d_hub_gokaku_added_address)}件): {', '.join(d_hub_gokaku_added_address)}\n"
            result_msg += f"\n合計 {updated_count} 箇所のセルを更新しました。"
        else:
            result_msg += "更新対象となるデータが見つかりませんでした。"

        all_results.append(result_msg)

    # 6. 変更を保存
    try:
        workbook.save(filepath)
        # 全シートの結果をまとめて表示
        final_message = f"ファイル '{os.path.basename(filepath)}' の処理が完了しました。\n\n"
        final_message += "\n\n---\n\n".join(all_results)
        messagebox.showinfo("処理完了", final_message)
    except Exception as e:
        error_detail = str(e)
        if "Permission denied" in error_detail:
            messagebox.showerror("保存エラー", f"ファイルの保存中にアクセスが拒否されました:\n{filepath}\n\nファイルがExcelなどの他のプログラムで開かれている可能性があります。\nファイルを閉じてから再度実行してください。\n\n詳細: {e}")
        else:
            messagebox.showerror("保存エラー", f"ファイルの保存中にエラーが発生しました:\n{filepath}\n\n詳細: {e}")

# --- メイン実行ブロック ---
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw() # メインウィンドウは表示しない

    # ファイル選択ダイアログを表示
    file_path = filedialog.askopenfilename(
        title="処理するExcelファイルを選択してください",
        filetypes=[("Excel ファイル", "*.xlsx *.xlsm")] # openpyxlはxlsx/xlsmをサポート
    )

    if file_path: # ファイルが選択された場合のみ処理を実行
        process_report_sheets(file_path)
    else:
        messagebox.showinfo("キャンセル", "ファイルが選択されなかったため、処理をキャンセルしました。")

    # tkinterのメインループを開始する必要はない (ダイアログとメッセージボックスのみ使用)
    # root.mainloop() は不要
